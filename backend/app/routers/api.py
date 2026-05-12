"""REST API 路由 - 供前端Web应用调用"""
import io
import os
import csv
import time
import re
import random
import logging
import json
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Query, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from app.services.bitable_client import bitable
from app.services.ocr_service import ocr_service
from app.services.ai_service import (
    ai_chat,
    ai_confirm,
    session_manager,
    init_ai_service,
)
from app.config import settings

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── 智能导入：同义词字典 ──────────────────────────────

FIELD_SYNONYMS = {
    "姓名": ["姓名", "名字", "name", "学员姓名", "学生姓名", "客户名", "会员名"],
    "微信昵称": ["微信昵称", "微信名", "昵称", "wechat", "vx名"],
    "电话": ["电话", "手机", "手机号", "联系方式", "tel", "phone", "联系电话", "移动电话"],
    "金额": ["金额", "价格", "费用", "付款金额", "实付", "收费", "多少钱", "付了多少钱", "售价", "实收金额", "实收"],
    "付款方式": ["付款方式", "支付方式", "支付", "付款", "收款方式"],
    "卡类型": ["卡类型", "类型", "卡类别"],
    "卡种名称": ["卡种名称", "卡种", "买的卡", "卡名", "课程名", "套餐名", "会员类型"],
    "总课时": ["总课时", "课时", "总次数", "次数", "课时数", "报卡类别", "课次"],
    "剩余课时": ["剩余课时", "剩余次数", "余课", "余次"],
    "激活日期": ["激活日期", "开始日期", "生效日期", "办卡日期", "购卡日期", "开始时间", "报卡日期"],
    "有效期至": ["有效期至", "到期日期", "有效期", "到期", "截止日期", "过期日期", "结束日期"],
    "渠道来源": ["渠道来源", "来源", "渠道", "获客渠道", "获客来源"],
    "备注": ["备注", "说明", "备注信息", "其他", "补充"],
    "性别": ["性别"],
    "出生日期": ["出生日期", "生日", "出生年月"],
    "学员类型": ["学员类型", "成人/少儿", "学员分类"],
    "卡类原价": ["卡类原价", "原价", "卡原价", "标准价"],
    "单课价": ["单课价", "单次价", "课时单价", "每课价格"],
}

# 模糊匹配黑名单：包含这些关键词的列名不应映射到7L字段
FUZZY_BLACKLIST = {
    "金额": ["提成", "退款", "原价", "单课价"],
    "剩余课时": ["提成", "比例", "发放"],
    "有效期至": ["顺延"],
    "出生日期": ["年龄"],  # 年龄是数字不是日期
    "学员类型": ["年龄"],  # 年龄不是学员类型
}

STANDARD_FIELDS = list(FIELD_SYNONYMS.keys())

logger = logging.getLogger(__name__)
router = APIRouter()


# ── 工具函数 ──────────────────────────────────────────

def _ts() -> int:
    return int(time.time() * 1000)


def _today_ms() -> int:
    now = datetime.now()
    return int(now.replace(hour=0, minute=0, second=0, microsecond=0).timestamp() * 1000)


def _days_later_ms(days: int) -> int:
    future = datetime.now() + timedelta(days=days)
    return int(future.replace(hour=23, minute=59, second=59).timestamp() * 1000)


def _month_start_ms() -> int:
    now = datetime.now()
    return int(now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).timestamp() * 1000)


def _fmt_date(ts_ms) -> str:
    if not ts_ms:
        return ""
    return datetime.fromtimestamp(ts_ms / 1000).strftime("%Y-%m-%d")


def _safe_float(val, default=0.0):
    try:
        return float(val) if val is not None else default
    except (ValueError, TypeError):
        return default


def _safe_int(val, default=0):
    try:
        return int(float(val)) if val is not None else default
    except (ValueError, TypeError):
        return default


def _record_to_student(r: dict) -> dict:
    """将Bitable记录转为前端友好的学员字典"""
    f = r.get("fields", {})
    return {
        "record_id": r.get("record_id", ""),
        "name": f.get("姓名", ""),
        "member_id": f.get("会员号", ""),
        "wechat": f.get("微信昵称", ""),
        "phone": f.get("电话", ""),
        "card_type": f.get("卡类型", ""),
        "card_name": f.get("卡种名称", ""),
        "amount": _safe_float(f.get("金额")),
        "total_hours": _safe_float(f.get("总课时")),
        "remaining_hours": _safe_float(f.get("剩余课时")),
        "card_status": f.get("卡状态", ""),
        "activate_date": _fmt_date(f.get("激活日期")),
        "expire_date": _fmt_date(f.get("有效期至")),
        "payment_method": f.get("付款方式", ""),
        "payment_date": _fmt_date(f.get("付款日期")),
        "channel": f.get("渠道来源", ""),
        "note": f.get("备注", ""),
    }


# ── Pydantic 模型 ─────────────────────────────────────

class RegisterRequest(BaseModel):
    name: str
    card_name: str = ""
    hours: Optional[float] = None  # 课次（可选，自由填写）
    amount: Optional[float] = None
    payment_method: str = ""
    phone: str = ""
    wechat: str = ""
    channel: str = ""
    note: str = ""


class DeductRequest(BaseModel):
    name: str
    deduct_count: int = 1
    teacher: str = ""
    dance_type: str = ""


class RenewRequest(BaseModel):
    name: str
    card_name: str = ""
    amount: Optional[float] = None
    payment_method: str = ""


class RefundRequest(BaseModel):
    record_id: Optional[str] = None


class UpdateStudentRequest(BaseModel):
    phone: Optional[str] = None
    wechat: Optional[str] = None
    note: Optional[str] = None


class AIChatRequest(BaseModel):
    message: str
    session_id: Optional[str] = None


class AIConfirmRequest(BaseModel):
    session_id: str
    confirmed: bool


# ── AI助手 ──────────────────────────────────────────

@router.post("/ai/chat")
async def ai_chat_endpoint(req: AIChatRequest):
    """AI对话（意图识别+参数提取+多轮对话）"""
    try:
        result = await ai_chat(req.message, req.session_id)
        return result
    except Exception as e:
        logger.error(f"AI对话失败: {e}", exc_info=True)
        return {
            "session_id": req.session_id or "",
            "reply": "AI暂时不可用，请用按钮操作",
            "state": "idle",
            "pending_action": None,
            "need_confirm": False,
        }


@router.post("/ai/confirm")
async def ai_confirm_endpoint(req: AIConfirmRequest):
    """确认执行AI操作"""
    try:
        result = await ai_confirm(req.session_id, req.confirmed)
        return result
    except Exception as e:
        logger.error(f"AI确认失败: {e}", exc_info=True)
        return {
            "session_id": req.session_id,
            "reply": "操作失败，请重试",
            "state": "idle",
            "pending_action": None,
            "need_confirm": False,
        }


# ── 学员列表 ──────────────────────────────────────────

@router.get("/students")
async def list_students(
    search: str = Query("", description="搜索关键词（姓名/会员号）"),
    card_status: str = Query("", description="卡状态筛选"),
    card_type: str = Query("", description="卡类型筛选"),
):
    """获取学员列表，支持搜索和筛选"""
    try:
        filter_parts = []
        if card_status:
            filter_parts.append(f'CurrentValue.[卡状态] = "{card_status}"')
        if card_type:
            filter_parts.append(f'CurrentValue.[卡类型] = "{card_type}"')

        filter_expr = " AND ".join(filter_parts) if filter_parts else None

        records = await bitable.list_records(
            settings.BITABLE_MAIN_APP_TOKEN,
            settings.BITABLE_MAIN_TABLE_ID,
            filter_expr=filter_expr,
        ) or []

        students = [_record_to_student(r) for r in records]

        # 客户端搜索（姓名/会员号模糊匹配）
        if search:
            search_lower = search.lower()
            students = [
                s for s in students
                if search_lower in (s.get("name") or "").lower()
                or search_lower in (s.get("member_id") or "").lower()
                or search_lower in (s.get("wechat") or "").lower()
                or search_lower in (s.get("phone") or "").lower()
            ]

        return {"students": students, "total": len(students)}

    except Exception as e:
        logger.error(f"获取学员列表失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ── 学员详情 ──────────────────────────────────────────

@router.get("/students/{name}")
async def get_student(name: str):
    """获取学员详情，包含上课记录和操作日志"""
    try:
        # 查找学员
        records = await bitable.find_member_by_name(name)
        if not records:
            raise HTTPException(status_code=404, detail=f"未找到学员「{name}」")

        student = _record_to_student(records[0])

        # 如果有多张卡，也返回
        all_cards = [_record_to_student(r) for r in records]
        student["all_cards"] = all_cards

        # 上课记录
        class_records_raw = await bitable.list_records(
            settings.BITABLE_CLASS_APP_TOKEN,
            settings.BITABLE_CLASS_TABLE_ID,
            filter_expr=f'CurrentValue.[学员姓名] = "{name}"',
        ) or []
        class_records = []
        for r in class_records_raw:
            f = r.get("fields", {})
            class_records.append({
                "date": _fmt_date(f.get("上课日期")),
                "teacher": f.get("老师", ""),
                "dance_type": f.get("舞种", ""),
                "deduct_count": _safe_int(f.get("扣课数")),
                "remaining_after": _safe_float(f.get("扣课后剩余")),
            })
        class_records.sort(key=lambda x: x["date"], reverse=True)
        student["class_records"] = class_records

        # 操作日志
        logs_raw = await bitable.get_recent_logs(student_name=name, limit=20)
        logs = []
        for r in logs_raw:
            f = r.get("fields", {})
            logs.append({
                "time": _fmt_date(f.get("操作时间")),
                "operator": f.get("操作人", ""),
                "type": f.get("操作类型", ""),
                "detail": f.get("变更详情", ""),
                "undone": bool(f.get("是否已撤销")),
            })
        student["logs"] = logs

        return student

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取学员详情失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ── 录入新学员 ────────────────────────────────────────

@router.post("/students")
async def register_student(req: RegisterRequest):
    """录入新学员（卡种自由化：卡种名称为可选标签，课次和金额自由填写）"""
    try:
        if not req.name:
            raise HTTPException(status_code=400, detail="姓名不能为空")

        # 课次和金额允许为0，但不能同时为0
        hours_val = req.hours if req.hours is not None else 0
        amount_val = req.amount if req.amount is not None else 0
        if hours_val == 0 and amount_val == 0:
            raise HTTPException(status_code=400, detail="课次和金额不能同时为0，请至少填写一项")

        # 检查是否已有有效卡
        existing = await bitable.find_member_by_name(req.name)
        if existing:
            active = [r for r in existing if r.get("fields", {}).get("卡状态") == "有效"]
            if active:
                raise HTTPException(status_code=400, detail=f"学员「{req.name}」已有有效卡，请先退卡或续费")

        # 查找卡种定价（仅作快捷模板参考，不强制匹配）
        card_info = None
        if req.card_name:
            pricing_records = await bitable.get_pricing_list()
            # 精确匹配
            for r in pricing_records:
                f = r.get("fields", {})
                if f.get("卡种名称") == req.card_name:
                    card_info = f
                    break
            # 模糊匹配
            if not card_info:
                for r in pricing_records:
                    f = r.get("fields", {})
                    name = f.get("卡种名称", "")
                    if name and req.card_name in name:
                        card_info = f
                        break

        # 构建记录
        fields = {
            "7L街舞工作室管理系统": req.name,
            "姓名": req.name,
        }

        if card_info:
            # 匹配到卡种定价表：用定价表作为默认值，但用户填写的值优先
            fields["卡类型"] = card_info.get("卡类型", "次卡")
            fields["卡种名称"] = card_info.get("卡种名称", req.card_name)
            # 课次：用户填了就用用户的，否则用定价表的
            if req.hours is not None:
                fields["总课时"] = float(req.hours)
                fields["剩余课时"] = float(req.hours)
            else:
                fields["总课时"] = float(card_info.get("课时数", 0))
                fields["剩余课时"] = float(card_info.get("课时数", 0))
            # 金额：用户填了就用用户的，否则用定价表的
            if req.amount is not None:
                fields["金额"] = float(req.amount)
            else:
                fields["金额"] = float(card_info.get("金额", 0))
            valid_days = int(float(card_info.get("有效期天", 30)))
            fields["激活日期"] = _ts()
            fields["有效期至"] = _days_later_ms(valid_days)
        else:
            # 未匹配到卡种定价表：自由填写
            fields["卡类型"] = "次卡"  # 默认次卡
            fields["卡种名称"] = req.card_name or "自定义"
            fields["总课时"] = float(hours_val)
            fields["剩余课时"] = float(hours_val)
            fields["金额"] = float(amount_val)
            fields["激活日期"] = _ts()
            fields["有效期至"] = _days_later_ms(30)

        if req.payment_method:
            fields["付款方式"] = req.payment_method
        if req.phone:
            fields["电话"] = req.phone
        if req.wechat:
            fields["微信昵称"] = req.wechat
        if req.channel:
            fields["渠道来源"] = req.channel
        if req.note:
            fields["备注"] = req.note

        fields["付款日期"] = _ts()
        fields["卡状态"] = "有效"

        member_id = f"7L{int(time.time())}{random.randint(1000, 9999)}"
        fields["会员号"] = member_id

        record = await bitable.create_record(
            settings.BITABLE_MAIN_APP_TOKEN,
            settings.BITABLE_MAIN_TABLE_ID,
            fields,
        )

        # 操作日志
        detail = f"录入卡种:{fields.get('卡种名称','')} 金额:{fields.get('金额',0)} 课时:{fields.get('总课时',0)}"
        await bitable.add_log("店长", "录入", req.name, detail)

        return {
            "success": True,
            "message": "录入成功",
            "member_id": member_id,
            "card_name": fields.get("卡种名称", ""),
            "total_hours": fields.get("总课时", 0),
            "amount": fields.get("金额", 0),
            "expire_date": _fmt_date(fields.get("有效期至")),
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"录入学员失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ── 扣课 ──────────────────────────────────────────────

@router.post("/deduct")
async def deduct_class(req: DeductRequest):
    """扣课签到"""
    try:
        if not req.name:
            raise HTTPException(status_code=400, detail="姓名不能为空")

        records = await bitable.find_member_by_name(req.name)
        if not records:
            raise HTTPException(status_code=404, detail=f"未找到学员「{req.name}」")

        active = [r for r in records if r.get("fields", {}).get("卡状态") == "有效"]
        if not active:
            raise HTTPException(status_code=400, detail=f"学员「{req.name}」没有有效卡")

        record = active[0]
        fields = record.get("fields", {})
        record_id = record.get("record_id")

        remaining = _safe_float(fields.get("剩余课时"))
        deduct = req.deduct_count or 1

        # 月卡不扣课时
        if fields.get("卡类型") == "月卡":
            expire = fields.get("有效期至", 0)
            if expire and expire < _ts():
                raise HTTPException(status_code=400, detail=f"学员「{req.name}」月卡已过期")

            await bitable.add_class_record({
                "7L-上课登记": f"{req.name}-{_fmt_date(_ts())}",
                "学员姓名": req.name,
                "上课日期": _ts(),
                "老师": req.teacher,
                "舞种": req.dance_type,
                "扣课数": 0,
                "扣课前剩余": remaining,
                "扣课后剩余": remaining,
                "关联卡号": fields.get("会员号", ""),
            })
            await bitable.add_log("店长", "扣课", req.name,
                                  f"月卡签到 老师:{req.teacher or '-'} 舞种:{req.dance_type or '-'}")
            return {"success": True, "message": "月卡签到成功", "remaining": remaining, "card_type": "月卡"}

        # 次卡/期卡/体验卡
        if remaining < deduct:
            raise HTTPException(status_code=400, detail=f"课时不足！剩余 {remaining} 次，需要扣 {deduct} 次")

        new_remaining = remaining - deduct
        await bitable.update_record(
            settings.BITABLE_MAIN_APP_TOKEN,
            settings.BITABLE_MAIN_TABLE_ID,
            record_id,
            {"剩余课时": new_remaining},
        )

        await bitable.add_class_record({
            "7L-上课登记": f"{req.name}-{_fmt_date(_ts())}",
            "学员姓名": req.name,
            "上课日期": _ts(),
            "老师": req.teacher,
            "舞种": req.dance_type,
            "扣课数": deduct,
            "扣课前剩余": remaining,
            "扣课后剩余": new_remaining,
            "关联卡号": fields.get("会员号", ""),
        })

        detail = f"扣{deduct}次 课前:{remaining} 课后:{new_remaining} 老师:{req.teacher or '-'}"
        await bitable.add_log("店长", "扣课", req.name, detail)

        warning = ""
        if new_remaining <= 2:
            warning = "课时即将用完，请提醒续费！"

        return {
            "success": True,
            "message": "扣课成功",
            "deducted": deduct,
            "remaining": new_remaining,
            "warning": warning,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"扣课失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ── 续费 ──────────────────────────────────────────────

@router.post("/renew")
async def renew_student(req: RenewRequest):
    """续费"""
    try:
        if not req.name:
            raise HTTPException(status_code=400, detail="姓名不能为空")

        records = await bitable.find_member_by_name(req.name)
        if not records:
            raise HTTPException(status_code=404, detail=f"未找到学员「{req.name}」")

        active = [r for r in records if r.get("fields", {}).get("卡状态") == "有效"]
        record = active[0] if active else records[0]
        fields = record.get("fields", {})
        record_id = record.get("record_id")

        # 查找新卡种
        card_info = None
        if req.card_name:
            pricing_records = await bitable.get_pricing_list()
            for r in pricing_records:
                f = r.get("fields", {})
                if f.get("卡种名称") == req.card_name:
                    card_info = f
                    break
            if not card_info:
                raise HTTPException(status_code=400, detail=f"未找到卡种「{req.card_name}」")

        old_remaining = _safe_float(fields.get("剩余课时"))

        if card_info:
            new_total = float(card_info.get("课时数", 0))
            new_remaining = old_remaining + new_total
            valid_days = int(float(card_info.get("有效期天", 30)))
            new_amount = req.amount if req.amount is not None else float(card_info.get("金额", 0))

            update_fields = {
                "卡类型": card_info.get("卡类型", fields.get("卡类型", "")),
                "卡种名称": card_info.get("卡种名称", req.card_name),
                "总课时": _safe_float(fields.get("总课时")) + new_total,
                "剩余课时": new_remaining,
                "金额": _safe_float(fields.get("金额")) + new_amount,
                "有效期至": _days_later_ms(valid_days),
                "卡状态": "有效",
            }
        else:
            # 无卡种信息，简单续费
            new_remaining = old_remaining
            new_amount = req.amount or 0
            update_fields = {
                "金额": _safe_float(fields.get("金额")) + new_amount,
                "卡状态": "有效",
            }

        if req.payment_method:
            update_fields["付款方式"] = req.payment_method

        await bitable.update_record(
            settings.BITABLE_MAIN_APP_TOKEN,
            settings.BITABLE_MAIN_TABLE_ID,
            record_id,
            update_fields,
        )

        detail = f"续费:{req.card_name} +{card_info.get('课时数', 0) if card_info else 0}课时 金额:{new_amount}"
        await bitable.add_log("店长", "续费", req.name, detail)

        return {
            "success": True,
            "message": "续费成功",
            "new_remaining": update_fields.get("剩余课时", new_remaining),
            "expire_date": _fmt_date(update_fields.get("有效期至")),
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"续费失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ── 退卡 ──────────────────────────────────────────────

@router.post("/students/{name:path}/refund")
async def refund_student(name: str, req: RefundRequest = None):
    """退卡"""
    try:
        records = await bitable.find_member_by_name(name)
        if not records:
            raise HTTPException(status_code=404, detail=f"未找到学员「{name}」")

        # 找到有效卡
        active = [r for r in records if r.get("fields", {}).get("卡状态") == "有效"]
        if not active:
            raise HTTPException(status_code=400, detail=f"学员「{name}」没有有效卡")

        record = active[0]
        record_id = req.record_id if req and req.record_id else record.get("record_id")

        # 更新卡状态
        await bitable.update_record(
            settings.BITABLE_MAIN_APP_TOKEN,
            settings.BITABLE_MAIN_TABLE_ID,
            record_id,
            {"卡状态": "已退卡"},
        )

        # 操作日志
        await bitable.add_log("店长", "退卡", name, f"退卡")

        return {"success": True, "message": "退卡成功"}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"退卡失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ── 更新学员信息 ──────────────────────────────────────────

@router.post("/students/{record_id}/update")
async def update_student(record_id: str, req: UpdateStudentRequest = None):
    """更新学员信息"""
    try:
        if not req:
            raise HTTPException(status_code=400, detail="请求体不能为空")

        fields = {}
        if req.phone is not None:
            fields["电话"] = req.phone
        if req.wechat is not None:
            fields["微信昵称"] = req.wechat
        if req.note is not None:
            fields["备注"] = req.note

        if not fields:
            raise HTTPException(status_code=400, detail="没有需要更新的字段")

        await bitable.update_record(
            settings.BITABLE_MAIN_APP_TOKEN,
            settings.BITABLE_MAIN_TABLE_ID,
            record_id,
            fields,
        )

        return {"success": True, "message": "更新成功"}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新学员信息失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ── 统计数据 ──────────────────────────────────────────

# ── stats缓存 ──────────────────────────────────────
_stats_cache = {"data": None, "ts": 0}
_STATS_CACHE_TTL = 120  # 2分钟


@router.get("/stats")
async def get_stats():
    """获取统计数据"""
    try:
        # 缓存检查
        now = time.time()
        if _stats_cache["data"] and (now - _stats_cache["ts"]) < _STATS_CACHE_TTL:
            return _stats_cache["data"]

        # 所有有效卡
        active = await bitable.get_active_members() or []

        # 所有课消记录（Bitable filter对DateTime字段不生效，改为内存过滤）
        all_class_records = await bitable.list_records(
            settings.BITABLE_CLASS_APP_TOKEN,
            settings.BITABLE_CLASS_TABLE_ID,
        ) or []

        # 本月上课记录
        month_start = _month_start_ms()
        class_records = [r for r in all_class_records if r.get("fields", {}).get("上课日期", 0) >= month_start]

        # 今日上课
        today_start = _today_ms()
        today_class = [r for r in class_records if r.get("fields", {}).get("上课日期", 0) >= today_start]

        # 本月课消金额（课消记录数 × 对应学员的单课价）
        # 构建学员名→单课价映射：单课价字段 → 金额/总课时 → 定价表卡种匹配 → 卡类型默认价
        # 1) 从定价表构建 卡种名称→单课价 映射
        pricing_records = await bitable.list_records(
            settings.BITABLE_PRICING_APP_TOKEN,
            settings.BITABLE_PRICING_TABLE_ID,
        ) or []
        pricing_price_map = {}  # 卡种名称 → 单课价
        for r in pricing_records:
            f = r.get("fields", {})
            card_name = f.get("卡种名称", "")
            amount = _safe_float(f.get("金额"))
            hours = _safe_float(f.get("课时数"))
            if card_name and amount > 0 and hours > 0:
                pricing_price_map[card_name] = round(amount / hours, 2)

        # 2) 从学员表构建 姓名→单课价
        all_main_records = await bitable.list_records(
            settings.BITABLE_MAIN_APP_TOKEN,
            settings.BITABLE_MAIN_TABLE_ID,
        ) or []
        student_price_map = {}
        for r in all_main_records:
            f = r.get("fields", {})
            name = f.get("姓名", "")
            if not name:
                continue
            price = _safe_float(f.get("单课价"))
            # 如果单课价为空，尝试从 金额/总课时 推算
            if price <= 0:
                amount = _safe_float(f.get("金额"))
                total_hours = _safe_float(f.get("总课时"))
                if amount > 0 and total_hours > 0:
                    price = round(amount / total_hours, 2)
            # 如果还是空，尝试从定价表匹配卡种名称
            if price <= 0:
                card_name = f.get("卡种名称", "")
                if card_name in pricing_price_map:
                    price = pricing_price_map[card_name]
            # 兜底：按卡类型给默认单课价
            if price <= 0:
                card_type = f.get("卡类型", "")
                default_prices = {"次卡": 80, "通卡": 50, "私教": 150, "体验卡": 100, "月卡": 40}
                price = default_prices.get(card_type, 80)
            if price > 0:
                student_price_map[name] = price

        # 本月课消金额
        month_class_revenue = 0
        for r in class_records:
            f = r.get("fields", {})
            sname = f.get("学员姓名", "")
            deduct = _safe_float(f.get("扣课数"), 1)
            price = student_price_map.get(sname, 0)
            if price > 0:
                month_class_revenue += price * deduct

        # 上月课消金额和次数（用于环比）
        last_month_start_dt = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0) - timedelta(days=1)
        last_month_start = int(last_month_start_dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0).timestamp() * 1000)
        last_month_end = month_start  # 上月末=本月初
        last_month_class_records = [
            r for r in all_class_records
            if last_month_start <= r.get("fields", {}).get("上课日期", 0) < last_month_end
        ]
        last_month_class_count = len(last_month_class_records)
        last_month_class_revenue = 0
        for r in last_month_class_records:
            f = r.get("fields", {})
            sname = f.get("学员姓名", "")
            deduct = _safe_float(f.get("扣课数"), 1)
            price = student_price_map.get(sname, 0)
            if price > 0:
                last_month_class_revenue += price * deduct

        # 环比变化
        mom_revenue_change = round(
            ((month_class_revenue - last_month_class_revenue) / max(last_month_class_revenue, 1)) * 100, 1
        ) if last_month_class_revenue > 0 else None
        mom_class_change = round(
            ((len(class_records) - last_month_class_count) / max(last_month_class_count, 1)) * 100, 1
        ) if last_month_class_count > 0 else None

        # 本月新学员（按付款日期）
        new_students_month = [
            r for r in all_main_records
            if r.get("fields", {}).get("付款日期", 0) and r["fields"]["付款日期"] >= month_start
        ]

        # 本月营收
        month_revenue = sum(_safe_float(r.get("fields", {}).get("金额")) for r in new_students_month)

        # 本月续费
        log_records = await bitable.list_records(
            settings.BITABLE_LOG_APP_TOKEN,
            settings.BITABLE_LOG_TABLE_ID,
            filter_expr=f'CurrentValue.[操作时间] >= {month_start}',
        ) or []
        renew_logs = [r for r in log_records if r.get("fields", {}).get("操作类型") == "续费"]
        renew_count = len(renew_logs)

        # 卡类型分布
        card_type_dist = {}
        for r in active:
            ct = r.get("fields", {}).get("卡类型", "未知")
            card_type_dist[ct] = card_type_dist.get(ct, 0) + 1

        # 即将过期（7天内）
        expiring_soon = []
        for r in active:
            f = r.get("fields", {})
            expire = f.get("有效期至", 0)
            if expire and expire < _days_later_ms(7):
                expiring_soon.append({
                    "name": f.get("姓名", ""),
                    "expire_date": _fmt_date(expire),
                    "card_name": f.get("卡种名称", ""),
                })

        # 课时不足（排除剩余=0的数据异常，只看1-2次的）
        low_hours = []
        for r in active:
            f = r.get("fields", {})
            remaining = _safe_float(f.get("剩余课时"))
            card_type = f.get("卡类型", "")
            if 0 < remaining <= 2 and card_type not in ("月卡", "通卡"):
                low_hours.append({
                    "name": f.get("姓名", ""),
                    "remaining": remaining,
                    "card_name": f.get("卡种名称", ""),
                })

        # 即将过生日（7天内）
        birthday_soon = []
        now = datetime.now()
        for r in active:
            f = r.get("fields", {})
            birth_ts = f.get("出生日期", 0)
            if birth_ts:
                try:
                    birth = datetime.fromtimestamp(birth_ts / 1000)
                    # 今年生日
                    this_bday = birth.replace(year=now.year)
                    if this_bday < now:
                        this_bday = birth.replace(year=now.year + 1)
                    days_until = (this_bday - now).days
                    if 0 <= days_until <= 7:
                        birthday_soon.append({
                            "name": f.get("姓名", ""),
                            "birthday": f"{birth.month}月{birth.day}日",
                            "days_until": days_until,
                        })
                except (ValueError, OSError):
                    pass

        # 续费率
        total_active = len(active)
        renew_rate = (renew_count / max(total_active, 1)) * 100

        result = {
            "today_classes": len(today_class),
            "month_revenue": month_revenue,
            "month_new_students": len(new_students_month),
            "month_class_count": len(class_records),
            "month_class_revenue": round(month_class_revenue, 2),
            "mom_revenue_change": mom_revenue_change,
            "mom_class_change": mom_class_change,
            "total_active": total_active,
            "renew_count": renew_count,
            "renew_rate": round(renew_rate, 1),
            "card_type_dist": card_type_dist,
            "expiring_soon": expiring_soon,
            "low_hours": low_hours,
            "reminder_count": len(expiring_soon) + len(low_hours),
            "birthday_soon": birthday_soon,
        }

        # 缓存
        _stats_cache["data"] = result
        _stats_cache["ts"] = time.time()
        return result

    except Exception as e:
        logger.error(f"获取统计失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ── 数据分析缓存 ──────────────────────────────────────
_analytics_cache = {"data": None, "ts": 0}
_ANALYTICS_CACHE_TTL = 300  # 5分钟


@router.get("/analytics")
async def get_analytics(
    start_date: Optional[str] = Query(None, description="开始日期 YYYY-MM"),
    end_date: Optional[str] = Query(None, description="结束日期 YYYY-MM"),
):
    """数据分析：月度/年度报卡金额、课消统计"""
    try:
        # 缓存检查
        now = time.time()
        if _analytics_cache["data"] and (now - _analytics_cache["ts"]) < _ANALYTICS_CACHE_TTL:
            data = _analytics_cache["data"]
        else:
            # 获取全部主表记录
            all_records = await bitable.list_records(
                settings.BITABLE_MAIN_APP_TOKEN,
                settings.BITABLE_MAIN_TABLE_ID,
            ) or []

            # 获取上课登记记录
            class_records = await bitable.list_records(
                settings.BITABLE_CLASS_APP_TOKEN,
                settings.BITABLE_CLASS_TABLE_ID,
            ) or []

            # ── 构建学员单课价映射（和stats接口一样的兜底逻辑） ──
            pricing_records = await bitable.list_records(
                settings.BITABLE_PRICING_APP_TOKEN,
                settings.BITABLE_PRICING_TABLE_ID,
            ) or []
            pricing_price_map = {}  # 卡种名称 → 单课价
            for r in pricing_records:
                f = r.get("fields", {})
                card_name = f.get("卡种名称", "")
                amount = _safe_float(f.get("金额"))
                hours = _safe_float(f.get("课时数"))
                if card_name and amount > 0 and hours > 0:
                    pricing_price_map[card_name] = round(amount / hours, 2)

            student_price_map = {}
            for r in all_records:
                f = r.get("fields", {})
                name = f.get("姓名", "")
                if not name:
                    continue
                price = _safe_float(f.get("单课价"))
                if price <= 0:
                    amount = _safe_float(f.get("金额"))
                    total_hours = _safe_float(f.get("总课时"))
                    if amount > 0 and total_hours > 0:
                        price = round(amount / total_hours, 2)
                if price <= 0:
                    card_name = f.get("卡种名称", "")
                    if card_name in pricing_price_map:
                        price = pricing_price_map[card_name]
                if price <= 0:
                    card_type = f.get("卡类型", "")
                    default_prices = {"次卡": 80, "通卡": 50, "私教": 150, "体验卡": 100, "月卡": 40}
                    price = default_prices.get(card_type, 80)
                if price > 0:
                    student_price_map[name] = price

            # 月度报卡金额（按报卡日期/激活日期分组）
            monthly_card = {}  # {"2022-08": {"amount": 12684, "count": 15}}
            for r in all_records:
                f = r.get("fields", {})
                amount = _safe_float(f.get("金额"))
                # 优先用激活日期（真实报卡日期），其次付款日期
                ts = f.get("激活日期") or f.get("付款日期")
                if not ts:
                    continue
                try:
                    dt = datetime.fromtimestamp(ts / 1000)
                    key = dt.strftime("%Y-%m")
                    if key not in monthly_card:
                        monthly_card[key] = {"amount": 0, "count": 0}
                    monthly_card[key]["amount"] += amount
                    monthly_card[key]["count"] += 1
                except (ValueError, OSError):
                    pass

            # 年度汇总
            yearly_card = {}
            for month_key, v in monthly_card.items():
                year = month_key[:4]
                if year not in yearly_card:
                    yearly_card[year] = {"amount": 0, "count": 0}
                yearly_card[year]["amount"] += v["amount"]
                yearly_card[year]["count"] += v["count"]

            # 月度课消统计
            monthly_class = {}
            has_class_data = len(class_records) > 0
            for r in class_records:
                f = r.get("fields", {})
                class_ts = f.get("上课日期", 0)
                deducted = _safe_float(f.get("扣课数", 1))
                if not class_ts:
                    continue
                try:
                    dt = datetime.fromtimestamp(class_ts / 1000)
                    key = dt.strftime("%Y-%m")
                    if key not in monthly_class:
                        monthly_class[key] = {"count": 0, "deducted": 0}
                    monthly_class[key]["count"] += 1
                    monthly_class[key]["deducted"] += deducted
                except (ValueError, OSError):
                    pass

            # 排序
            monthly_card_sorted = dict(sorted(monthly_card.items()))
            yearly_card_sorted = dict(sorted(yearly_card.items()))
            monthly_class_sorted = dict(sorted(monthly_class.items()))

            # 同比/环比
            months_list = sorted(monthly_card.keys())
            yoy = {}
            mom = {}
            for m in months_list:
                # 同比：去年同月
                year, mon = m.split("-")
                last_year_m = f"{int(year)-1}-{mon}"
                if last_year_m in monthly_card:
                    yoy[m] = round((monthly_card[m]["amount"] - monthly_card[last_year_m]["amount"]) / max(monthly_card[last_year_m]["amount"], 1) * 100, 1)
                # 环比：上月
                mi = months_list.index(m)
                if mi > 0:
                    prev_m = months_list[mi - 1]
                    mom[m] = round((monthly_card[m]["amount"] - monthly_card[prev_m]["amount"]) / max(monthly_card[prev_m]["amount"], 1) * 100, 1)

            # ── 新增：年度课消金额 ──
            yearly_class_revenue = {}
            for r in class_records:
                f = r.get("fields", {})
                class_ts = f.get("上课日期", 0)
                if not class_ts:
                    continue
                try:
                    dt = datetime.fromtimestamp(class_ts / 1000)
                    year_key = str(dt.year)
                except (ValueError, OSError):
                    continue
                sname = f.get("学员姓名", "")
                deduct = _safe_float(f.get("扣课数"), 1)
                price = student_price_map.get(sname, 0)
                if year_key not in yearly_class_revenue:
                    yearly_class_revenue[year_key] = 0.0
                if price > 0:
                    yearly_class_revenue[year_key] += price * deduct
            # round
            yearly_class_revenue = {k: round(v, 2) for k, v in yearly_class_revenue.items()}

            # ── 新增：老师开课统计 ──
            teacher_stats_map = {}  # {"小七": {"count": 35, "dances": set(), "revenue": 0.0}}
            for r in class_records:
                f = r.get("fields", {})
                teacher = f.get("老师", "") or "未知"
                dance = f.get("舞种", "") or ""
                sname = f.get("学员姓名", "")
                deduct = _safe_float(f.get("扣课数"), 1)
                price = student_price_map.get(sname, 0)
                if teacher not in teacher_stats_map:
                    teacher_stats_map[teacher] = {"count": 0, "dances": set(), "revenue": 0.0}
                teacher_stats_map[teacher]["count"] += 1
                if dance:
                    teacher_stats_map[teacher]["dances"].add(dance)
                if price > 0:
                    teacher_stats_map[teacher]["revenue"] += price * deduct
            total_class_count = sum(v["count"] for v in teacher_stats_map.values())
            teacher_stats = sorted([
                {
                    "name": name,
                    "count": v["count"],
                    "pct": round(v["count"] / max(total_class_count, 1) * 100, 1),
                    "dances": sorted(list(v["dances"])),
                    "revenue": round(v["revenue"], 2),
                }
                for name, v in teacher_stats_map.items()
            ], key=lambda x: x["count"], reverse=True)

            # ── 新增：舞种热度统计（含趋势） ──
            dance_stats_map = {}  # {"HipHop": 30}
            dance_monthly_map = {}  # {"HipHop": {"2026-03": 5, "2026-04": 8}}
            for r in class_records:
                f = r.get("fields", {})
                dance = f.get("舞种", "") or "未知"
                dance_stats_map[dance] = dance_stats_map.get(dance, 0) + 1
                class_ts = f.get("上课日期", 0)
                if class_ts:
                    try:
                        dt = datetime.fromtimestamp(class_ts / 1000)
                        mkey = dt.strftime("%Y-%m")
                        if dance not in dance_monthly_map:
                            dance_monthly_map[dance] = {}
                        dance_monthly_map[dance][mkey] = dance_monthly_map[dance].get(mkey, 0) + 1
                    except (ValueError, OSError):
                        pass
            dance_total = sum(dance_stats_map.values())
            # 本月和上月
            now_dt = datetime.now()
            this_month_key = now_dt.strftime("%Y-%m")
            last_month_dt = now_dt.replace(day=1) - timedelta(days=1)
            last_month_key = last_month_dt.strftime("%Y-%m")
            dance_stats = sorted([
                {
                    "name": name,
                    "count": count,
                    "pct": round(count / max(dance_total, 1) * 100, 1),
                    "trend": "↑" if dance_monthly_map.get(name, {}).get(this_month_key, 0) >= dance_monthly_map.get(name, {}).get(last_month_key, 0) else "↓",
                }
                for name, count in dance_stats_map.items()
            ], key=lambda x: x["count"], reverse=True)

            # ── 新增：月度课消次数 ──
            monthly_class_count = {}
            for r in class_records:
                f = r.get("fields", {})
                class_ts = f.get("上课日期", 0)
                if not class_ts:
                    continue
                try:
                    dt = datetime.fromtimestamp(class_ts / 1000)
                    key = dt.strftime("%Y-%m")
                except (ValueError, OSError):
                    continue
                monthly_class_count[key] = monthly_class_count.get(key, 0) + 1
            monthly_class_count_sorted = dict(sorted(monthly_class_count.items()))

            # ── 新增：学员健康度 ──
            now_dt = datetime.now()
            this_month_start = int(now_dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0).timestamp() * 1000)
            thirty_days_ago = int((now_dt - timedelta(days=30)).timestamp() * 1000)

            # 活跃学员：30天内有课消记录
            active_class_names = set()
            for r in class_records:
                f = r.get("fields", {})
                class_ts = f.get("上课日期", 0)
                if class_ts and class_ts >= thirty_days_ago:
                    active_class_names.add(f.get("学员姓名", ""))

            # 有效卡学员
            active_members = [r for r in all_records if r.get("fields", {}).get("卡状态") == "有效"]
            active_member_names = set(r.get("fields", {}).get("姓名", "") for r in active_members)

            active_count = len(active_class_names & active_member_names)
            silent_count = len(active_member_names - active_class_names)

            # 本月新学员（付款日期在本月）
            new_this_month = sum(
                1 for r in all_records
                if r.get("fields", {}).get("付款日期", 0) and r["fields"]["付款日期"] >= this_month_start
            )

            # 流失学员：卡状态为已过期且有效期在本月之前
            lost_count = sum(
                1 for r in all_records
                if r.get("fields", {}).get("卡状态") == "已过期"
                and r.get("fields", {}).get("有效期至", 0)
                and r["fields"]["有效期至"] < this_month_start
                and r["fields"]["有效期至"] >= int((now_dt.replace(day=1) - timedelta(days=90)).timestamp() * 1000)  # 近3个月流失
            )

            # 平均每学员月上课次数
            total_active_members = max(len(active_member_names), 1)
            this_month_class_count = sum(
                1 for r in class_records
                if r.get("fields", {}).get("上课日期", 0) and r["fields"]["上课日期"] >= this_month_start
            )
            avg_monthly_classes = round(this_month_class_count / total_active_members, 1)

            # 续费率趋势（近3个月）
            renew_trend = []
            for i in range(3):
                m_dt = (now_dt.replace(day=1) - timedelta(days=1) * 31 * i).replace(day=1)
                m_start = int(m_dt.timestamp() * 1000)
                next_m = (m_dt + timedelta(days=32)).replace(day=1)
                m_end = int(next_m.timestamp() * 1000)
                m_key = m_dt.strftime("%Y-%m")
                # 该月续费次数
                m_renew_count = sum(
                    1 for r in class_records
                    if r.get("fields", {}).get("上课日期", 0) and m_start <= r["fields"]["上课日期"] < m_end
                )
                m_active_count = sum(
                    1 for r in all_records
                    if r.get("fields", {}).get("卡状态") == "有效"
                    and r.get("fields", {}).get("激活日期", 0)
                    and r["fields"]["激活日期"] < m_end
                )
                m_rate = round(m_renew_count / max(m_active_count, 1) * 100, 1)
                renew_trend.append({"month": m_key, "rate": m_rate})
            renew_trend.sort(key=lambda x: x["month"])

            student_health = {
                "active_count": active_count,
                "silent_count": silent_count,
                "new_this_month": new_this_month,
                "lost_count": lost_count,
                "avg_monthly_classes": avg_monthly_classes,
                "renew_trend": renew_trend,
            }

            # ── 新增：卡类分析 ──
            # 各卡类型分布
            card_type_dist_map = {}
            for r in active_members:
                ct = r.get("fields", {}).get("卡类型", "") or "未知"
                card_type_dist_map[ct] = card_type_dist_map.get(ct, 0) + 1
            total_active_cards = max(sum(card_type_dist_map.values()), 1)
            card_type_distribution = sorted([
                {"type": ct, "count": cnt, "pct": round(cnt / total_active_cards * 100, 1)}
                for ct, cnt in card_type_dist_map.items()
            ], key=lambda x: x["count"], reverse=True)

            # 各卡类型课消贡献
            card_type_class_map = {}  # 卡类型 → 课消次数
            member_card_type_map = {}  # 学员名 → 卡类型
            for r in all_records:
                f = r.get("fields", {})
                name = f.get("姓名", "")
                ct = f.get("卡类型", "") or "未知"
                if name and ct:
                    member_card_type_map[name] = ct
            for r in class_records:
                f = r.get("fields", {})
                sname = f.get("学员姓名", "")
                ct = member_card_type_map.get(sname, "未知")
                card_type_class_map[ct] = card_type_class_map.get(ct, 0) + 1
            total_class_by_card = max(sum(card_type_class_map.values()), 1)
            class_contribution = sorted([
                {"type": ct, "count": cnt, "pct": round(cnt / total_class_by_card * 100, 1)}
                for ct, cnt in card_type_class_map.items()
            ], key=lambda x: x["count"], reverse=True)

            # 即将过期卡数（7天内）
            week_later_ms = _days_later_ms(7)
            expiring_soon_count = sum(
                1 for r in active_members
                if r.get("fields", {}).get("有效期至", 0) and r["fields"]["有效期至"] < week_later_ms
            )

            card_type_stats = {
                "distribution": card_type_distribution,
                "class_contribution": class_contribution,
                "expiring_soon": expiring_soon_count,
            }

            data = {
                "monthly_card": monthly_card_sorted,
                "yearly_card": yearly_card_sorted,
                "monthly_class": monthly_class_sorted,
                "has_class_data": has_class_data,
                "yoy": yoy,
                "mom": mom,
                "total_records": len(all_records),
                "total_class_records": len(class_records),
                "yearly_class_revenue": yearly_class_revenue,
                "teacher_stats": teacher_stats,
                "dance_stats": dance_stats,
                "monthly_class_count": monthly_class_count_sorted,
                "student_health": student_health,
                "card_type_stats": card_type_stats,
            }

            _analytics_cache["data"] = data
            _analytics_cache["ts"] = now

        # 日期范围过滤
        if start_date or end_date:
            filtered_monthly = {}
            for k, v in data["monthly_card"].items():
                if start_date and k < start_date:
                    continue
                if end_date and k > end_date:
                    continue
                filtered_monthly[k] = v
            data = {**data, "monthly_card": filtered_monthly}

        return data

    except Exception as e:
        logger.error(f"获取分析数据失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
async def get_recent_operations():
    """获取最近10条操作记录"""
    try:
        records = await bitable.get_recent_logs(limit=10)
        operations = []
        now_ts = _ts()
        seven_days_ms = 7 * 24 * 60 * 60 * 1000
        for r in records:
            f = r.get("fields", {})
            op_time = f.get("操作时间", 0)
            undone = bool(f.get("是否已撤销"))
            operations.append({
                "record_id": r.get("record_id", ""),
                "type": f.get("操作类型", ""),
                "student_name": f.get("学员姓名", ""),
                "detail": f.get("变更详情", ""),
                "time": _fmt_date(op_time),
                "time_ms": op_time,
                "undone": undone,
                "can_undo": (not undone) and (now_ts - op_time < seven_days_ms),
            })
        return {"operations": operations}
    except Exception as e:
        logger.error(f"获取最近操作失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/undo/{record_id}")
async def undo_operation(record_id: str):
    """撤销操作"""
    try:
        # 获取操作记录
        try:
            log_record = await bitable.get_record(
                settings.BITABLE_LOG_APP_TOKEN,
                settings.BITABLE_LOG_TABLE_ID,
                record_id,
            )
        except Exception:
            raise HTTPException(status_code=404, detail="操作记录不存在")

        f = log_record.get("fields", {})

        # 检查是否已撤销
        if f.get("是否已撤销"):
            raise HTTPException(status_code=400, detail="该操作已被撤销")

        # 检查7天窗口期
        op_time = f.get("操作时间", 0)
        if _ts() - op_time > 7 * 24 * 60 * 60 * 1000:
            raise HTTPException(status_code=400, detail="已超过7天撤销窗口期")

        op_type = f.get("操作类型", "")
        student_name = f.get("学员姓名", "")

        # 不允许撤销"撤销"类型的操作
        if op_type == "撤销":
            raise HTTPException(status_code=400, detail="撤销操作不可再次撤销")

        # 根据操作类型执行撤销
        if op_type == "扣课":
            # 撤销扣课：恢复课时
            detail = f.get("变更详情", "")
            # 解析 detail: "扣1次 课前:16 课后:15 老师:xx"
            import re
            match = re.search(r"扣(\d+)次\s+课前:([\d.]+)\s+课后:([\d.]+)", detail)
            if match:
                deduct_count = int(match.group(1))
                before_remaining = float(match.group(2))
                # 找到学员的有效卡
                student_records = await bitable.find_member_by_name(student_name)
                if student_records:
                    active = [r for r in student_records if r.get("fields", {}).get("卡状态") == "有效"]
                    if active:
                        await bitable.update_record(
                            settings.BITABLE_MAIN_APP_TOKEN,
                            settings.BITABLE_MAIN_TABLE_ID,
                            active[0]["record_id"],
                            {"剩余课时": before_remaining},
                        )
            # 标记日志为已撤销
            await bitable.update_record(
                settings.BITABLE_LOG_APP_TOKEN,
                settings.BITABLE_LOG_TABLE_ID,
                record_id,
                {"是否已撤销": True},
            )
            await bitable.add_log("店长", "撤销", student_name, f"撤销扣课: {detail}")

        elif op_type == "录入":
            # 撤销录入：将卡状态改为已退卡
            student_records = await bitable.find_member_by_name(student_name)
            if student_records:
                for r in student_records:
                    if r.get("fields", {}).get("卡状态") == "有效":
                        await bitable.update_record(
                            settings.BITABLE_MAIN_APP_TOKEN,
                            settings.BITABLE_MAIN_TABLE_ID,
                            r["record_id"],
                            {"卡状态": "已退卡"},
                        )
                        break
            await bitable.update_record(
                settings.BITABLE_LOG_APP_TOKEN,
                settings.BITABLE_LOG_TABLE_ID,
                record_id,
                {"是否已撤销": True},
            )
            await bitable.add_log("店长", "撤销", student_name, f"撤销录入")

        elif op_type == "续费":
            # 撤销续费：简单标记日志，不回退金额（续费回退逻辑复杂，需手动处理）
            await bitable.update_record(
                settings.BITABLE_LOG_APP_TOKEN,
                settings.BITABLE_LOG_TABLE_ID,
                record_id,
                {"是否已撤销": True},
            )
            await bitable.add_log("店长", "撤销", student_name, f"撤销续费: {f.get('变更详情', '')}")

        else:
            # 其他操作类型，只标记撤销
            await bitable.update_record(
                settings.BITABLE_LOG_APP_TOKEN,
                settings.BITABLE_LOG_TABLE_ID,
                record_id,
                {"是否已撤销": True},
            )
            await bitable.add_log("店长", "撤销", student_name, f"撤销{op_type}")

        return {"success": True, "message": f"已撤销{op_type}操作"}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"撤销操作失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ── 提醒列表 ──────────────────────────────────────────

@router.get("/reminders")
async def get_reminders():
    """获取提醒列表"""
    try:
        active = await bitable.get_active_members() or []

        expiring = []
        inactive = []
        low_hours = []

        now_ts = _ts()
        week_later = _days_later_ms(7)

        for r in active:
            f = r.get("fields", {})
            name = f.get("姓名", "")
            card_name = f.get("卡种名称", "")
            remaining = _safe_float(f.get("剩余课时"))
            expire = f.get("有效期至", 0)
            card_type = f.get("卡类型", "")

            # 即将过期（7天内）
            if expire and expire < week_later:
                expiring.append({
                    "name": name,
                    "type": "即将过期",
                    "card_name": card_name,
                    "expire_date": _fmt_date(expire),
                    "detail": f"有效期至 {_fmt_date(expire)}",
                })

            # 久未上课（检查上课记录）
            # 简化：用激活日期判断，如果超过30天没上课记录
            # 这里先跳过，需要查上课登记表，太慢

            # 课时不足
            if card_type != "月卡" and remaining <= 2:
                low_hours.append({
                    "name": name,
                    "type": "课时不足",
                    "card_name": card_name,
                    "remaining": remaining,
                    "detail": f"剩余 {remaining} 次",
                })

        # 查久未上课（最近30天没有上课记录的有效学员）
        month_ago = _days_later_ms(-30)
        class_records = await bitable.list_records(
            settings.BITABLE_CLASS_APP_TOKEN,
            settings.BITABLE_CLASS_TABLE_ID,
            filter_expr=f'CurrentValue.[上课日期] >= {month_ago}',
        ) or []
        recent_class_names = set()
        for r in class_records:
            recent_class_names.add(r.get("fields", {}).get("学员姓名", ""))

        for r in active:
            f = r.get("fields", {})
            name = f.get("姓名", "")
            if name not in recent_class_names:
                activate = f.get("激活日期", 0)
                if activate and activate < month_ago:
                    inactive.append({
                        "name": name,
                        "type": "久未上课",
                        "card_name": f.get("卡种名称", ""),
                        "detail": "超过30天未上课",
                    })

        return {
            "expiring": expiring,
            "low_hours": low_hours,
            "inactive": inactive,
            "total": len(expiring) + len(low_hours) + len(inactive),
        }

    except Exception as e:
        logger.error(f"获取提醒失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ── OCR 截图识别 ────────────────────────────────────────

@router.post("/ocr")
async def ocr_recognize(file: UploadFile = File(...)):
    """接收截图图片，调用百度OCR识别，返回结构化学员列表"""
    try:
        # 读取图片
        image_bytes = await file.read()
        if not image_bytes:
            raise HTTPException(status_code=400, detail="图片为空")

        # 限制文件大小 10MB
        if len(image_bytes) > 10 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="图片太大，请压缩后重试（最大10MB）")

        # 调用百度OCR识别
        lines = await ocr_service.recognize_image(image_bytes)

        if not lines:
            return {
                "success": True,
                "raw_text": [],
                "students": [],
                "message": "未识别到文字内容",
            }

        # 解析接龙文本
        students = ocr_service.parse_jielong_text(lines)

        return {
            "success": True,
            "raw_text": lines,
            "students": students,
            "message": f"识别到 {len(lines)} 行文字，{len(students)} 位学员",
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"OCR识别失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"OCR识别失败: {str(e)}")


# ── OCR 批量录入 ────────────────────────────────────────

class OCRBatchRegisterRequest(BaseModel):
    students: list[dict]  # [{"name": "张三", "note": "体验课", "card_name": "", "amount": 0}]


@router.post("/ocr/batch-register")
async def ocr_batch_register(req: OCRBatchRegisterRequest):
    """批量录入OCR识别出的学员"""
    results = []
    for s in req.students:
        name = s.get("name", "").strip()
        if not name:
            results.append({"name": "", "success": False, "error": "姓名为空"})
            continue

        try:
            # 复用录入逻辑
            register_req = RegisterRequest(
                name=name,
                card_name=s.get("card_name", ""),
                amount=s.get("amount"),
                payment_method=s.get("payment_method", "微信"),
                phone=s.get("phone", ""),
                wechat=s.get("wechat", ""),
                channel=s.get("channel", ""),
                note=s.get("note", ""),
            )
            result = await register_student(register_req)
            results.append({"name": name, "success": True, "member_id": result.get("member_id", "")})
        except HTTPException as e:
            results.append({"name": name, "success": False, "error": e.detail})
        except Exception as e:
            results.append({"name": name, "success": False, "error": str(e)})

    success_count = sum(1 for r in results if r["success"])
    return {
        "success": True,
        "total": len(results),
        "success_count": success_count,
        "fail_count": len(results) - success_count,
        "results": results,
    }


# ── 卡种定价 ──────────────────────────────────────────

class CardPriceCreateRequest(BaseModel):
    name: str
    card_type: str = "次卡"
    price: Optional[float] = 0
    hours: Optional[int] = 0
    valid_days: Optional[int] = 30
    note: str = ""


class CardPriceUpdateRequest(BaseModel):
    name: Optional[str] = None
    card_type: Optional[str] = None
    price: Optional[float] = None
    hours: Optional[int] = None
    valid_days: Optional[int] = None
    note: Optional[str] = None


@router.get("/card-prices")
async def get_card_prices():
    """获取卡种定价列表"""
    try:
        records = await bitable.get_pricing_list()
        prices = []
        for r in records:
            f = r.get("fields", {})
            name = f.get("卡种名称", "")
            if not name:
                continue
            prices.append({
                "record_id": r.get("record_id", ""),
                "name": name,
                "card_type": f.get("卡类型", ""),
                "price": _safe_float(f.get("金额")),
                "hours": _safe_int(f.get("课时数")),
                "valid_days": _safe_int(f.get("有效期天")),
                "note": f.get("说明", ""),
            })
        return {"prices": prices}

    except Exception as e:
        logger.error(f"获取卡种定价失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/card-prices")
async def create_card_price(req: CardPriceCreateRequest):
    """新增卡种定价"""
    try:
        if not req.name:
            raise HTTPException(status_code=400, detail="卡种名称不能为空")

        # 检查是否重名
        existing = await bitable.get_pricing_list()
        for r in existing:
            if r.get("fields", {}).get("卡种名称") == req.name:
                raise HTTPException(status_code=400, detail=f"卡种「{req.name}」已存在")

        fields = {
            "卡种名称": req.name,
            "卡类型": req.card_type,
            "金额": req.price or 0,
            "课时数": req.hours or 0,
            "有效期天": req.valid_days or 30,
        }
        if req.note:
            fields["说明"] = req.note

        record = await bitable.create_record(
            settings.BITABLE_PRICING_APP_TOKEN,
            settings.BITABLE_PRICING_TABLE_ID,
            fields,
        )

        # 刷新缓存
        await bitable.get_pricing_list()

        return {
            "success": True,
            "message": f"卡种「{req.name}」创建成功",
            "record_id": record.get("record_id", ""),
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"新增卡种定价失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/card-prices/{record_id}")
async def update_card_price(record_id: str, req: CardPriceUpdateRequest):
    """修改卡种定价"""
    try:
        fields = {}
        if req.name is not None:
            fields["卡种名称"] = req.name
        if req.card_type is not None:
            fields["卡类型"] = req.card_type
        if req.price is not None:
            fields["金额"] = req.price
        if req.hours is not None:
            fields["课时数"] = req.hours
        if req.valid_days is not None:
            fields["有效期天"] = req.valid_days
        if req.note is not None:
            fields["说明"] = req.note

        if not fields:
            raise HTTPException(status_code=400, detail="没有需要更新的字段")

        # 检查重名（排除自身）
        if req.name is not None:
            existing = await bitable.get_pricing_list()
            for r in existing:
                if (r.get("fields", {}).get("卡种名称") == req.name
                        and r.get("record_id") != record_id):
                    raise HTTPException(status_code=400, detail=f"卡种「{req.name}」已存在")

        await bitable.update_record(
            settings.BITABLE_PRICING_APP_TOKEN,
            settings.BITABLE_PRICING_TABLE_ID,
            record_id,
            fields,
        )

        # 刷新缓存
        await bitable.get_pricing_list()

        return {
            "success": True,
            "message": "修改成功",
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"修改卡种定价失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/card-prices/{record_id}")
async def delete_card_price(record_id: str):
    """删除卡种定价"""
    try:
        await bitable.delete_record(
            settings.BITABLE_PRICING_APP_TOKEN,
            settings.BITABLE_PRICING_TABLE_ID,
            record_id,
        )

        # 刷新缓存
        await bitable.get_pricing_list()

        return {
            "success": True,
            "message": "删除成功",
        }

    except Exception as e:
        logger.error(f"删除卡种定价失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ── 老师管理 ──────────────────────────────────────────

class TeacherCreateRequest(BaseModel):
    name: str
    dances: list[str] = []
    phone: str = ""
    join_date: Optional[str] = None  # YYYY-MM-DD
    status: str = "在教"
    note: str = ""


class TeacherUpdateRequest(BaseModel):
    name: Optional[str] = None
    dances: Optional[list[str]] = None
    phone: Optional[str] = None
    join_date: Optional[str] = None
    status: Optional[str] = None
    note: Optional[str] = None


def _record_to_teacher(r: dict) -> dict:
    """将Bitable记录转为前端友好的老师字典"""
    f = r.get("fields", {})
    # 舞种可能是 list（MultiSelect）或 str
    dances = f.get("舞种", [])
    if isinstance(dances, str):
        dances = [d.strip() for d in dances.split(",") if d.strip()]
    return {
        "record_id": r.get("record_id", ""),
        "name": f.get("姓名", ""),
        "dances": dances or [],
        "phone": f.get("手机号", ""),
        "join_date": _fmt_date(f.get("入职日期")),
        "status": f.get("状态", "在教"),
        "note": f.get("备注", ""),
    }


@router.get("/teachers")
async def list_teachers():
    """获取老师列表"""
    try:
        records = await bitable.get_teachers()
        teachers = [_record_to_teacher(r) for r in records]
        # 在教排前面
        teachers.sort(key=lambda t: (0 if t["status"] == "在教" else 1, t["name"]))
        return {"teachers": teachers}
    except Exception as e:
        logger.error(f"获取老师列表失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/teachers")
async def create_teacher(req: TeacherCreateRequest):
    """新增老师"""
    try:
        if not req.name:
            raise HTTPException(status_code=400, detail="姓名不能为空")

        if not settings.BITABLE_TEACHER_APP_TOKEN or not settings.BITABLE_TEACHER_TABLE_ID:
            raise HTTPException(status_code=500, detail="老师表未配置，请先运行 seed 脚本")

        # 检查重名
        existing = await bitable.get_teachers()
        for r in existing:
            if r.get("fields", {}).get("姓名") == req.name:
                raise HTTPException(status_code=400, detail=f"老师「{req.name}」已存在")

        fields = {
            "姓名": req.name,
            "舞种": req.dances,
            "状态": req.status or "在教",
        }
        if req.phone:
            fields["手机号"] = req.phone
        if req.join_date:
            try:
                dt = datetime.strptime(req.join_date, "%Y-%m-%d")
                fields["入职日期"] = int(dt.timestamp() * 1000)
            except ValueError:
                pass
        if req.note:
            fields["备注"] = req.note

        record = await bitable.create_teacher(fields)
        return {
            "success": True,
            "message": f"老师「{req.name}」创建成功",
            "record_id": record.get("record_id", ""),
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"新增老师失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/teachers/{record_id}")
async def update_teacher(record_id: str, req: TeacherUpdateRequest):
    """编辑老师信息"""
    try:
        fields = {}
        if req.name is not None:
            fields["姓名"] = req.name
        if req.dances is not None:
            fields["舞种"] = req.dances
        if req.phone is not None:
            fields["手机号"] = req.phone
        if req.join_date is not None:
            try:
                dt = datetime.strptime(req.join_date, "%Y-%m-%d")
                fields["入职日期"] = int(dt.timestamp() * 1000)
            except ValueError:
                pass
        if req.status is not None:
            fields["状态"] = req.status
        if req.note is not None:
            fields["备注"] = req.note

        if not fields:
            raise HTTPException(status_code=400, detail="没有需要更新的字段")

        await bitable.update_teacher(record_id, fields)
        return {"success": True, "message": "修改成功"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"编辑老师失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/teachers/{record_id}")
async def delete_teacher(record_id: str):
    """停用老师（改状态为停用，不做物理删除）"""
    try:
        await bitable.update_teacher(record_id, {"状态": "停用"})
        return {"success": True, "message": "已停用"}
    except Exception as e:
        logger.error(f"停用老师失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ── 智能导入：列名映射API ──────────────────────────────

def _rule_match_column(col_name: str) -> tuple[str | None, str, str]:
    """规则层匹配：同义词字典，返回 (mapped_to, confidence, method)"""
    # 清理列名：去掉换行符、多余空格
    col_clean = re.sub(r'\s+', '', col_name.strip())
    col_lower = col_clean.lower()
    
    # 精确匹配
    for field, synonyms in FIELD_SYNONYMS.items():
        for syn in synonyms:
            syn_clean = re.sub(r'\s+', '', syn.strip())
            if col_lower == syn_clean.lower():
                return field, "high", "rule"
    
    # 模糊匹配：列名包含同义词（带黑名单过滤）
    for field, synonyms in FIELD_SYNONYMS.items():
        for syn in synonyms:
            syn_clean = re.sub(r'\s+', '', syn.strip())
            if syn_clean.lower() in col_lower or col_lower in syn_clean.lower():
                # 黑名单检查
                blacklist = FUZZY_BLACKLIST.get(field, [])
                if any(bl.lower() in col_lower for bl in blacklist):
                    continue  # 跳过黑名单匹配
                return field, "medium", "rule"
    
    return None, "none", "skip"


def _desensitize_sample(value: str, field_hint: str | None = None) -> str:
    """样本数据脱敏：手机号只发前3位+***"""
    if not value:
        return ""
    # 手机号脱敏：11位数字
    if re.match(r'^1\d{10}$', value.strip()):
        return value[:3] + "***"
    return value


async def _ai_match_columns(unmatched: list[dict]) -> list[dict]:
    """AI层映射：通过OpenClaw Gateway调用AI对未匹配列名进行智能映射"""
    from app.services.ai_service import call_llm

    # 构建prompt
    columns_info = []
    for item in unmatched:
        col_info = {"original": item["original"], "samples": item["samples"]}
        columns_info.append(col_info)

    prompt = f"""你是数据导入助手。请将以下Excel列名映射到7L街舞工作室的标准字段。

7L标准字段列表：{json.dumps(STANDARD_FIELDS, ensure_ascii=False)}

待映射的列：
{json.dumps(columns_info, ensure_ascii=False)}

映射规则：
1. 如果列名能准确对应到某个标准字段，返回该字段名
2. 如果列名含义模糊（如"金额/课时"可能对应多个字段），选择最可能的那个，置信度标为medium
3. 如果列名无法映射到任何标准字段（如"教练偏好"、"第几次来店"），返回null，不要强行匹配
4. 注意区分：金额=实收金额（元），总课时=课次数，剩余课时=剩余课次数
5. 以下列名不应映射：提成相关列（提成发放比例/金额/客服）、顺延相关列（顺延后的截止日期）、原价列（卡类原价已单独映射）、年龄列（年龄是数字不是出生日期）

请返回JSON数组，每个元素包含：
- original: 原始列名
- mapped_to: 映射到的标准字段名（无法映射则为null）
- confidence: 置信度 "high"/"medium"/"low"

只返回JSON数组，不要其他文字。"""

    try:
        result = await call_llm([{"role": "user", "content": prompt}], timeout=10.0)
        if not result:
            logger.info("Gateway AI映射返回空，跳过AI映射")
            return []
        # 提取JSON
        match = re.search(r'\[.*\]', result, re.DOTALL)
        if match:
            return json.loads(match.group())
    except Exception as e:
        logger.warning(f"AI映射失败（降级为手动）: {e}")
    return []


@router.post("/import/smart-map")
async def smart_map_columns(file: UploadFile = File(...)):
    """智能列名映射：上传Excel/CSV，返回列名到7L标准字段的映射建议"""
    try:
        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail="文件为空")

        if len(content) > 5 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="文件太大，请控制在5MB以内")

        # 解析文件
        filename = file.filename or ""
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        try:
            if ext in ("xlsx", "xls"):
                rows = _parse_excel_bytes(content)
            else:
                rows = _parse_csv_bytes(content)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))

        if not rows:
            raise HTTPException(status_code=400, detail="文件为空或没有有效数据行")

        # 获取列名
        columns = list(rows[0].keys())

        # 提取每列前3行样本
        samples_map = {}
        for col in columns:
            samples = []
            for row in rows[:3]:
                val = row.get(col, "")
                samples.append(_desensitize_sample(str(val), col))
            samples_map[col] = samples

        # 第一层：规则匹配
        results = []
        unmatched = []
        for col in columns:
            mapped_to, confidence, method = _rule_match_column(col)
            if mapped_to:
                results.append({
                    "original": col,
                    "mapped_to": mapped_to,
                    "confidence": confidence,
                    "method": method,
                    "samples": samples_map[col],
                })
            else:
                unmatched.append({
                    "original": col,
                    "samples": samples_map[col],
                })

        # 第二层：AI映射（仅对规则未命中的列）
        if unmatched:
            ai_results = await _ai_match_columns(unmatched)
            ai_map = {r.get("original"): r for r in ai_results}

            for item in unmatched:
                ai_result = ai_map.get(item["original"])
                if ai_result and ai_result.get("mapped_to"):
                    results.append({
                        "original": item["original"],
                        "mapped_to": ai_result["mapped_to"],
                        "confidence": ai_result.get("confidence", "low"),
                        "method": "ai",
                        "samples": item["samples"],
                    })
                else:
                    results.append({
                        "original": item["original"],
                        "mapped_to": None,
                        "confidence": "none",
                        "method": "skip",
                        "samples": item["samples"],
                    })

        # 返回前5行预览数据（供前端展示）
        preview_rows = rows[:5]

        return {
            "columns": results,
            "standard_fields": STANDARD_FIELDS,
            "preview_rows": preview_rows,
            "total_rows": len(rows),
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"智能映射失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"映射失败: {str(e)}")


# ── 批量导入 ──────────────────────────────────────────

# CSV 模板列定义
IMPORT_CSV_COLUMNS = [
    "姓名", "微信昵称", "电话", "金额", "付款方式",
    "卡类型", "卡种名称", "总课时", "剩余课时",
    "激活日期", "有效期至", "渠道来源", "备注",
]


@router.get("/import/template")
async def download_import_template():
    """下载CSV导入模板"""
    output = io.StringIO()
    writer = csv.writer(output)
    # 写表头
    writer.writerow(IMPORT_CSV_COLUMNS)
    # 写示例行
    writer.writerow([
        "张三", "张三微信", "13800138000", "980", "微信",
        "次卡", "次卡·16次", "16", "16",
        "2026-04-30", "2026-08-28", "转介绍", "",
    ])
    content = output.getvalue()
    # BOM for Excel UTF-8 compatibility
    bom_content = "\ufeff" + content
    return StreamingResponse(
        io.BytesIO(bom_content.encode("utf-8-sig")),
        media_type="text/csv; charset=utf-8-sig",
        headers={
            "Content-Disposition": "attachment; filename=7L_import_template.csv",
        },
    )


def _parse_csv_bytes(content: bytes) -> list[dict]:
    """解析CSV字节内容，支持UTF-8和GBK编码"""
    # 尝试UTF-8
    text = None
    for encoding in ["utf-8-sig", "utf-8", "gbk", "gb18030"]:
        try:
            text = content.decode(encoding)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    if text is None:
        raise ValueError("无法识别文件编码，请使用UTF-8或GBK编码")

    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for row in reader:
        # 跳过全空行
        if all(not v or not v.strip() for v in row.values()):
            continue
        rows.append({k.strip(): (v.strip() if v else "") for k, v in row.items()})
    return rows


def _parse_excel_bytes(content: bytes) -> list[dict]:
    """解析Excel字节内容（.xlsx/.xls）"""
    if not HAS_OPENPYXL:
        raise ValueError("服务器未安装openpyxl，暂不支持Excel文件，请转换为CSV后上传")
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    
    # 智能查找表头行：找到包含"姓名"的行作为表头
    headers = None
    for row in rows_iter:
        row_strs = [str(v).strip() if v else "" for v in row]
        if "姓名" in row_strs or "学员姓名" in row_strs:
            headers = row_strs
            break
    
    if not headers:
        # 没找到姓名列，回退到第一行当表头
        wb.close()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = [str(h).strip() if h else "" for h in next(rows_iter)]
        except StopIteration:
            wb.close()
            return []
    
    rows = []
    for row in rows_iter:
        # 跳过全空行
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        d = {}
        for h, v in zip(headers, row):
            if h:
                val = str(v).strip() if v is not None else ""
                d[h] = val
        rows.append(d)
    wb.close()
    return rows


def _validate_import_row(row: dict, row_num: int) -> tuple[dict, str | None]:
    """校验单行导入数据，返回(fields_dict, error_msg)"""
    # 列名别名映射
    name = row.get("姓名", "") or row.get("学员姓名", "") or row.get("名字", "")
    name = name.strip()
    if not name:
        return {}, f"第{row_num}行：姓名不能为空"

    # 金额校验
    amount_str = (row.get("金额", "") or row.get("实收金额（元）", "") or row.get("实收金额", "")).strip()
    amount = 0.0
    if amount_str:
        try:
            amount = float(amount_str)
        except ValueError:
            return {}, f"第{row_num}行：金额格式错误「{amount_str}」"

    # 课时校验
    total_hours_str = (row.get("总课时", "") or row.get("报卡类别", "") or row.get("课时", "")).strip()
    total_hours = 0.0
    if total_hours_str:
        try:
            total_hours = float(total_hours_str)
        except ValueError:
            total_hours = 0.0  # 月卡等非数字课次设为0

    remaining_hours_str = row.get("剩余课时", "").strip()
    remaining_hours = total_hours  # 默认剩余=总课时
    if remaining_hours_str:
        try:
            remaining_hours = float(remaining_hours_str)
        except ValueError:
            return {}, f"第{row_num}行：剩余课时格式错误「{remaining_hours_str}」"

    # 日期校验
    def _parse_date(val: str) -> int | None:
        if not val:
            return None
        for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d", "%Y-%m-%d %H:%M:%S"]:
            try:
                dt = datetime.strptime(val, fmt)
                return int(dt.timestamp() * 1000)
            except ValueError:
                continue
        return None

    activate_date_str = (row.get("激活日期", "") or row.get("报卡日期", "")).strip()
    activate_date = _parse_date(activate_date_str)
    if activate_date_str and not activate_date:
        return {}, f"第{row_num}行：激活日期格式错误「{activate_date_str}」，请用YYYY-MM-DD"

    expire_date_str = (row.get("有效期至", "") or row.get("截止日期", "")).strip()
    expire_date = _parse_date(expire_date_str)
    if expire_date_str and not expire_date:
        return {}, f"第{row_num}行：有效期至格式错误「{expire_date_str}」，请用YYYY-MM-DD"

    # 构建fields
    fields = {
        "7L街舞工作室管理系统": name,
        "姓名": name,
        "金额": amount,
        "总课时": total_hours,
        "剩余课时": remaining_hours,
        "卡状态": "有效",
        "会员号": f"7L{int(time.time())}{random.randint(1000, 9999)}",
        "付款日期": _ts(),
    }

    # 可选字段（支持别名映射）
    wechat = (row.get("微信昵称", "") or row.get("微信名", "")).strip()
    if wechat:
        fields["微信昵称"] = wechat
    phone = (row.get("电话", "") or row.get("手机号", "")).strip()
    if phone:
        fields["电话"] = phone
    payment = (row.get("付款方式", "") or row.get("收款方式", "")).strip()
    if payment:
        fields["付款方式"] = payment
    card_type = (row.get("卡类型", "") or row.get("卡种", "")).strip()
    if card_type:
        fields["卡类型"] = card_type
    card_name = (row.get("卡种名称", "") or row.get("卡种", "")).strip()
    if card_name:
        fields["卡种名称"] = card_name
    channel = (row.get("渠道来源", "") or row.get("获客来源", "") or row.get("来源", "")).strip()
    if channel:
        fields["渠道来源"] = channel
    if row.get("备注", "").strip():
        fields["备注"] = row["备注"].strip()

    # 新增字段
    gender = (row.get("性别", "")).strip()
    if gender:
        fields["性别"] = gender
    student_type = (row.get("学员类型", "") or row.get("成人/少儿", "")).strip()
    if student_type:
        fields["学员类型"] = student_type
    card_original_price_str = (row.get("卡类原价", "") or row.get("卡类原价（元）", "")).strip()
    if card_original_price_str:
        try:
            fields["卡类原价"] = float(card_original_price_str)
        except ValueError:
            pass
    single_price_str = (row.get("单课价", "")).strip()
    if single_price_str:
        try:
            fields["单课价"] = float(single_price_str)
        except ValueError:
            pass
    birth_date_str = (row.get("出生日期", "") or row.get("生日", "")).strip()
    if birth_date_str:
        bd = _parse_date(birth_date_str)
        if bd:
            fields["出生日期"] = bd

    # 日期
    if activate_date:
        fields["激活日期"] = activate_date
    else:
        fields["激活日期"] = _ts()

    if expire_date:
        fields["有效期至"] = expire_date
    else:
        fields["有效期至"] = _days_later_ms(30)

    return fields, None


@router.post("/import")
async def import_students(file: UploadFile = File(...)):
    """批量导入学员（支持CSV和Excel文件）"""
    try:
        # 读取文件
        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail="文件为空")

        # 限制文件大小 5MB
        if len(content) > 5 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="文件太大，请控制在5MB以内")

        # 根据文件后缀选择解析方式
        filename = file.filename or ""
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        try:
            if ext in ("xlsx", "xls"):
                rows = _parse_excel_bytes(content)
            else:
                rows = _parse_csv_bytes(content)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))

        if not rows:
            raise HTTPException(status_code=400, detail="文件为空或没有有效数据行")

        # 获取卡种定价（用于自动匹配）
        pricing_records = await bitable.get_pricing_list()
        pricing_map = {}
        for r in pricing_records:
            f = r.get("fields", {})
            pname = f.get("卡种名称", "")
            if pname:
                pricing_map[pname] = f

        # 逐行校验 + 自动匹配卡种
        validated = []  # [(fields, row_num, name)]
        errors = []
        for i, row in enumerate(rows, start=2):  # 第1行是表头，数据从第2行开始
            fields, err = _validate_import_row(row, i)
            if err:
                errors.append({"row": i, "name": row.get("姓名", ""), "error": err})
                continue

            # 自动匹配卡种定价
            card_name = fields.get("卡种名称", "")
            if card_name and card_name in pricing_map:
                card_info = pricing_map[card_name]
                # 如果CSV没填总课时，用定价表的
                if not fields.get("总课时"):
                    fields["总课时"] = float(card_info.get("课时数", 0))
                    fields["剩余课时"] = float(card_info.get("课时数", 0))
                # 如果CSV没填金额，用定价表的
                if not fields.get("金额"):
                    fields["金额"] = float(card_info.get("金额", 0))
                # 如果CSV没填卡类型，用定价表的
                if not fields.get("卡类型") or fields.get("卡类型") == "次卡":
                    fields["卡类型"] = card_info.get("卡类型", "次卡")
                # 如果CSV没填有效期，用定价表的天数
                expire_str = row.get("有效期至", "").strip()
                if not expire_str:
                    valid_days = int(float(card_info.get("有效期天", 30)))
                    fields["有效期至"] = _days_later_ms(valid_days)

            validated.append((fields, i, fields["姓名"]))

        if not validated:
            return {
                "success": False,
                "total": len(rows),
                "success_count": 0,
                "fail_count": len(errors),
                "results": [],
                "errors": errors,
                "message": f"所有 {len(rows)} 行数据校验失败",
            }

        # 批量写入Bitable（分批，每批最多10条）
        success_results = []
        fail_results = []
        batch_size = 10

        for batch_start in range(0, len(validated), batch_size):
            batch = validated[batch_start:batch_start + batch_size]
            batch_fields = [item[0] for item in batch]

            try:
                created = await bitable.batch_create_records(
                    settings.BITABLE_MAIN_APP_TOKEN,
                    settings.BITABLE_MAIN_TABLE_ID,
                    batch_fields,
                )
                for j, item in enumerate(batch):
                    fields, row_num, name = item
                    if j < len(created):
                        member_id = fields.get("会员号", "")
                        success_results.append({
                            "row": row_num,
                            "name": name,
                            "success": True,
                            "member_id": member_id,
                        })
                    else:
                        fail_results.append({
                            "row": row_num,
                            "name": name,
                            "error": "写入Bitable失败",
                        })
            except Exception as e:
                logger.error(f"批量写入失败: {e}", exc_info=True)
                # 整批失败，逐条标记
                for item in batch:
                    _, row_num, name = item
                    fail_results.append({
                        "row": row_num,
                        "name": name,
                        "error": f"写入失败: {str(e)[:100]}",
                    })

        # 操作日志
        if success_results:
            detail = f"批量导入{len(success_results)}位学员"
            await bitable.add_log("店长", "批量导入", "-", detail)

        all_errors = errors + [{"row": r["row"], "name": r["name"], "error": r["error"]} for r in fail_results]

        return {
            "success": True,
            "total": len(rows),
            "success_count": len(success_results),
            "fail_count": len(all_errors),
            "results": success_results + fail_results,
            "errors": all_errors,
            "message": f"导入完成：成功 {len(success_results)} 人" + (f"，失败 {len(all_errors)} 人" if all_errors else ""),
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"批量导入失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")


@router.post("/import/with-mapping")
async def import_with_mapping(
    file: UploadFile = File(...),
    mapping: str = Form(...),
):
    """带映射关系的导入：先smart-map确认映射，再按映射导入"""
    try:
        col_mapping = json.loads(mapping)  # {"原始列名": "7L标准字段名"}

        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail="文件为空")

        filename = file.filename or ""
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        try:
            if ext in ("xlsx", "xls"):
                rows = _parse_excel_bytes(content)
            else:
                rows = _parse_csv_bytes(content)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))

        if not rows:
            raise HTTPException(status_code=400, detail="文件为空或没有有效数据行")

        # 按映射关系转换每行数据
        mapped_rows = []
        for row in rows:
            new_row = {}
            for original_col, target_field in col_mapping.items():
                if target_field and row.get(original_col):
                    new_row[target_field] = row[original_col]
            mapped_rows.append(new_row)

        # 获取卡种定价
        pricing_records = await bitable.get_pricing_list()
        pricing_map = {}
        for r in pricing_records:
            f = r.get("fields", {})
            pname = f.get("卡种名称", "")
            if pname:
                pricing_map[pname] = f

        # 逐行校验 + 导入
        validated = []
        errors = []
        for i, row in enumerate(mapped_rows, start=2):
            fields, err = _validate_import_row(row, i)
            if err:
                errors.append({"row": i, "name": row.get("姓名", ""), "error": err})
                continue

            card_name = fields.get("卡种名称", "")
            if card_name and card_name in pricing_map:
                card_info = pricing_map[card_name]
                if not fields.get("总课时"):
                    fields["总课时"] = float(card_info.get("课时数", 0))
                    fields["剩余课时"] = float(card_info.get("课时数", 0))
                if not fields.get("金额"):
                    fields["金额"] = float(card_info.get("金额", 0))
                if not fields.get("卡类型") or fields.get("卡类型") == "次卡":
                    fields["卡类型"] = card_info.get("卡类型", "次卡")
                expire_str = row.get("有效期至", "").strip()
                if not expire_str:
                    valid_days = int(float(card_info.get("有效期天", 30)))
                    fields["有效期至"] = _days_later_ms(valid_days)

            validated.append((fields, i, fields["姓名"]))

        if not validated:
            return {
                "success": False,
                "total": len(mapped_rows),
                "success_count": 0,
                "fail_count": len(errors),
                "results": [],
                "errors": errors,
                "message": f"所有 {len(mapped_rows)} 行数据校验失败",
            }

        success_results = []
        fail_results = []
        batch_size = 10
        for batch_start in range(0, len(validated), batch_size):
            batch = validated[batch_start:batch_start + batch_size]
            batch_fields = [item[0] for item in batch]
            try:
                created = await bitable.batch_create_records(
                    settings.BITABLE_MAIN_APP_TOKEN,
                    settings.BITABLE_MAIN_TABLE_ID,
                    batch_fields,
                )
                for j, item in enumerate(batch):
                    fields, row_num, name = item
                    if j < len(created):
                        member_id = fields.get("会员号", "")
                        success_results.append({"row": row_num, "name": name, "success": True, "member_id": member_id})
                    else:
                        fail_results.append({"row": row_num, "name": name, "success": False, "error": "写入失败"})
            except Exception as e:
                for item in batch:
                    fail_results.append({"row": item[1], "name": item[2], "success": False, "error": str(e)})

        all_errors = errors + fail_results
        return {
            "success": True,
            "total": len(mapped_rows),
            "success_count": len(success_results),
            "fail_count": len(all_errors),
            "results": success_results + fail_results,
            "errors": all_errors,
            "message": f"导入完成：成功 {len(success_results)} 人" + (f"，失败 {len(all_errors)} 人" if all_errors else ""),
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"带映射导入失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")


# ── 老师管理 ──────────────────────────────────
class TeacherCreate(BaseModel):
    name: str
    dances: list[str] = []
    phone: str = ""
    join_date: Optional[str] = None  # YYYY-MM-DD
    status: str = "在教"
    note: str = ""

class TeacherUpdate(BaseModel):
    name: Optional[str] = None
    dances: Optional[list[str]] = None
    phone: Optional[str] = None
    join_date: Optional[str] = None
    status: Optional[str] = None
    note: Optional[str] = None

@router.get("/teachers")
async def list_teachers():
    """获取老师列表"""
    app_token = settings.BITABLE_TEACHER_APP_TOKEN
    table_id = settings.BITABLE_TEACHER_TABLE_ID
    if not app_token or not table_id:
        return []
    try:
        records = await bitable.list_records(app_token, table_id, page_size=100)
        result = []
        for r in records:
            f = r.get("fields", {})
            join_ms = f.get("入职日期", 0)
            join_str = ""
            if join_ms:
                try:
                    join_str = datetime.fromtimestamp(join_ms / 1000).strftime("%Y-%m-%d")
                except:
                    pass
            result.append({
                "record_id": r.get("record_id", ""),
                "name": f.get("姓名", ""),
                "dances": f.get("舞种", []),
                "phone": f.get("手机号", ""),
                "join_date": join_str,
                "status": f.get("状态", "在教"),
                "note": f.get("备注", ""),
            })
        return result
    except Exception as e:
        logger.error(f"获取老师列表失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/teachers")
async def create_teacher(req: TeacherCreate):
    """新增老师"""
    app_token = settings.BITABLE_TEACHER_APP_TOKEN
    table_id = settings.BITABLE_TEACHER_TABLE_ID
    if not app_token or not table_id:
        raise HTTPException(status_code=400, detail="老师表未配置")
    try:
        fields = {
            "姓名": req.name,
            "舞种": req.dances,
            "手机号": req.phone,
            "状态": req.status,
            "备注": req.note,
        }
        if req.join_date:
            try:
                dt = datetime.strptime(req.join_date, "%Y-%m-%d")
                fields["入职日期"] = int(dt.timestamp() * 1000)
            except:
                pass
        record = await bitable.create_record(app_token, table_id, fields)
        return {"success": True, "record_id": record.get("record_id", ""), "message": f"老师 {req.name} 添加成功"}
    except Exception as e:
        logger.error(f"新增老师失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.put("/teachers/{record_id}")
async def update_teacher(record_id: str, req: TeacherUpdate):
    """编辑老师"""
    app_token = settings.BITABLE_TEACHER_APP_TOKEN
    table_id = settings.BITABLE_TEACHER_TABLE_ID
    if not app_token or not table_id:
        raise HTTPException(status_code=400, detail="老师表未配置")
    try:
        fields = {}
        if req.name is not None:
            fields["姓名"] = req.name
        if req.dances is not None:
            fields["舞种"] = req.dances
        if req.phone is not None:
            fields["手机号"] = req.phone
        if req.status is not None:
            fields["状态"] = req.status
        if req.note is not None:
            fields["备注"] = req.note
        if req.join_date is not None:
            try:
                dt = datetime.strptime(req.join_date, "%Y-%m-%d")
                fields["入职日期"] = int(dt.timestamp() * 1000)
            except:
                pass
        if not fields:
            raise HTTPException(status_code=400, detail="没有要更新的字段")
        await bitable.update_record(app_token, table_id, record_id, fields)
        return {"success": True, "message": "更新成功"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新老师失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/teachers/{record_id}")
async def delete_teacher(record_id: str):
    """停用老师（改状态为停用，不物理删除）"""
    app_token = settings.BITABLE_TEACHER_APP_TOKEN
    table_id = settings.BITABLE_TEACHER_TABLE_ID
    if not app_token or not table_id:
        raise HTTPException(status_code=400, detail="老师表未配置")
    try:
        await bitable.update_record(app_token, table_id, record_id, {"状态": "停用"})
        return {"success": True, "message": "老师已停用"}
    except Exception as e:
        logger.error(f"停用老师失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ── 数据清空 ──────────────────────────────────
@router.post("/clear-data/{data_type}")
async def clear_data(data_type: str):
    """清空指定类型的数据（危险操作）"""
    if data_type == "class_records":
        app_token = settings.BITABLE_CLASS_APP_TOKEN
        table_id = settings.BITABLE_CLASS_TABLE_ID
    elif data_type == "teachers":
        app_token = settings.BITABLE_TEACHER_APP_TOKEN
        table_id = settings.BITABLE_TEACHER_TABLE_ID
    elif data_type == "students":
        app_token = settings.BITABLE_MAIN_APP_TOKEN
        table_id = settings.BITABLE_MAIN_TABLE_ID
    elif data_type == "logs":
        app_token = settings.BITABLE_LOG_APP_TOKEN
        table_id = settings.BITABLE_LOG_TABLE_ID
    else:
        raise HTTPException(status_code=400, detail=f"不支持的数据类型: {data_type}")

    if not app_token or not table_id:
        raise HTTPException(status_code=400, detail="表未配置")

    try:
        # 获取所有记录
        records = await bitable.list_records(app_token, table_id, page_size=500)
        if not records:
            return {"deleted_count": 0, "message": "没有数据需要清空"}

        # 批量删除（每批10条）
        deleted = 0
        for i in range(0, len(records), 10):
            batch = records[i:i+10]
            for r in batch:
                try:
                    await bitable.delete_record(app_token, table_id, r["record_id"])
                    deleted += 1
                except Exception as e:
                    logger.warning(f"删除记录失败 {r.get('record_id')}: {e}")

        # 清除缓存
        _analytics_cache.clear()
        _stats_cache.clear()

        return {"deleted_count": deleted, "message": f"已清空 {deleted} 条记录"}
    except Exception as e:
        logger.error(f"清空数据失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))
